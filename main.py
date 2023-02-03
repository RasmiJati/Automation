import openpyxl

inventory_file = openpyxl.load_workbook("Inventory.xlsx", )
product_list = inventory_file["Sheet1"]   # sheet1 from excel file
products_per_supplier = {}  # empty dictionary
total_value_per_supplier = {}
products_under_10_inventory = {}
for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    total_price = product_list.cell(product_row, 5)

    # calculation number pf products per supplier
    # products_per_supplier["key"] = "value"
    if supplier_name in products_per_supplier:
        # current_num_product = products_per_supplier[supplier_name]
        current_num_product = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_product + 1
    else:
        products_per_supplier[supplier_name] = 1

    #  calculation total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # logic products for inventory under 10
    if inventory < 10:
        products_under_10_inventory[product_num] = inventory

    # add a column and insert total value of product
    total_price.value = inventory * price

print(products_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inventory)

inventory_file.save("inventory_with_total_value.xls")
